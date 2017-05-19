import os
import re

import pkg_resources

try:
    pkg_resources.get_distribution('pandas')
except pkg_resources.DistributionNotFound:
    phjPandasPresent = False
    print("Error: Pandas package not available.")
else:
    phjPandasPresent = True
    import pandas as pd


try:
    pkg_resources.get_distribution('numpy')
except pkg_resources.DistributionNotFound:
    phjNumpyPresent = False
    print("Error: Numpy package not available.")
else:
    phjNumpyPresent = True
    import numpy as np


try:
    pkg_resources.get_distribution('openpyxl')
except pkg_resources.DistributionNotFound:
    phjOpenpyxlPresent = False
    print("Error: Openpyxl package not available.")
else:
    phjOpenpyxlPresent = True
    import openpyxl



#######################################################
# Read data from a named cell range in Excel workbook #
#######################################################
def phjReadDataFromExcelNamedCellRange(phjExcelPathAndFileName = None,
                                       phjExcelCellRangeName = None,
                                       phjDatetimeFormat = "%Y-%m-%d %H:%M:%S",
                                       phjMissingValue = "missing",
                                       phjHeaderRow = False,
                                       phjPrintResults = False):
    
    phjAllowedAttempts = 3
    
    # This function reads data from a named cell range in an Excel spreadsheet (.xlsx).
    # Input phjExcelPathAndFileName (including path to file) and phjExcelCellRangeName.
    # Also input phjDatetimeFormat and phjPrintResults (default = False).
    # If the function can find the correct cell range, it reads the data and returns it in
    # the form of a pandas DataFrame. If the cell range can't be found, it returns None.
    
    # Since Openpyxl 2.4.1, the get_named_range() function was deprecated. This function was used by
    # the original version of phjReadDataFromExcelNamedCellRange(). However, a new method was required
    # for versions of Openpyxl 2.4.1 and later.
    if pkg_resources.get_distribution("openpyxl").version < '2.4.1':
    
        # OPENPYXL VERSIONS PRIOR TO 2.4.1
        # ================================
        # Find workbook and load cell range
        phjTempWorkbook, phjTempCellRange = phjLoadExcelWorkbook(phjExcelPathAndFileName = phjExcelPathAndFileName,
                                                                 phjExcelCellRangeName = phjExcelCellRangeName,
                                                                 phjMaxAttempts = phjAllowedAttempts,
                                                                 phjPrintResults = False)
        
        
        if phjTempCellRange is None:
            
            if phjPrintResults == True:
                print('\nData could not be retrieved from Excel workbook.\n')
            
            phjTempDF = None
        
        else:
            # Get name of worksheet from cell range instance
            # The cellrange.destinations is given as a list of tuples of the form:
            #
            #     [(<ReadOnlyWorksheet "Sheet1">, '$A$1:$D$100')]
            #
            # The first item in the first tuple is found using cellrange.destinations[0][0].
            # This needs to be converted to a string and a regular expression used to find the
            # text between the double quotation marks. The result is returned as a list (although,
            # in this case, there is only a single item). The first item is returned and
            # converted to a string. This can then be used to get a Worksheet instance.
            phjTempWorksheetName = str(re.findall(r'\"(.+?)\"',str(phjTempCellRange.destinations[0][0]))[0])
            phjTempWorksheet = phjTempWorkbook[phjTempWorksheetName]
            
            
            if phjPrintResults == True:
                print("\nList of iterable properties:")
                print(dir(phjTempCellRange))                        # Get list of all iterable properties of object
                
                print("\nList of tuples of named ranges and cell ranges")
                print(phjTempCellRange.destinations)                    # This gives a list containing tuples of worksheet names and cell ranges 
                
                print("\nFirst tuple of named ranges and cell range")
                print(phjTempCellRange.destinations[0])                # Give first tuple
                
                print("\nCell range")
                print(phjTempCellRange.destinations[0][1])            # Gives element [1] of tuple [0] i.e. the cell range
                
                print("\nLocal Sheet ID")
                print(phjTempCellRange.destinations[0][0])            # Gives element [1] of tuple [0] i.e. the cell range
                
                
            # Define temporary list to store data...
            phjTempImportedData = []
            
            # Step through each row in cell range.
            # (N.B. Cells returned by iter_rows() are not regular openpyxl.cell.cell.Cell but openpyxl.cell.read_only.ReadOnlyCell.)
            for phjTempRow in phjTempWorksheet.iter_rows(phjTempCellRange.destinations[0][1]):
                
                phjTempRowData = phjReadCells(phjTempRow,
                                              phjDatetimeFormat = phjDatetimeFormat,
                                              phjMissingValue = phjMissingValue)
                
                phjTempImportedData.append(tuple(phjTempRowData))
            
            # Deal with header row
            phjTempVariableNames = phjDealWithHeaderRow(phjTempImportedData,
                                                        phjHeaderRow = phjHeaderRow,
                                                        phjPrintResults = phjPrintResults)
            
            # Convert dataset to pandas DataFrame.
            # Each column now headed with original column headers as seen in Excel file
            # if header row present or with generic labels of 'var1', 'var2', etc. if no
            # header row present.
            phjTempDF = pd.DataFrame(phjTempImportedData, columns=phjTempVariableNames)
            
            
    else:
        
        # OPENPYXL VERSION 2.4.1 AND LATER
        # ================================
        # The original function details were modified to ensure compatibility with Openpyxl 2.4.1 and later. To
        # avoid modifying the original functions and risking breaking existing code, he modifications were added
        # as a separate section that would run only if the installed version of Openpyxl was 2.4.1 or later. Some
        # minor changes were made to the original functions but these should not create any compatibility issues.
        phjTempWorkbook = phjLoadExcelWorkbookNewOpenpyxl(phjExcelPathAndFileName = phjExcelPathAndFileName,
                                                          phjExcelCellRangeName = phjExcelCellRangeName,
                                                          phjMaxAttempts = phjAllowedAttempts,
                                                          phjPrintResults = False)
        
        if phjTempWorkbook is None:
            
            if phjPrintResults == True:
                print('\nData could not be retrieved from Excel workbook.\n')
            
            phjTempDF = None
        
        else:
            phjTempCellRange = phjLoadExcelNamedCellRangeNewOpenpyxl(phjExcelFileNameOnly = os.path.basename(phjExcelPathAndFileName),
                                                                     phjExcelWorkbook = phjTempWorkbook,
                                                                     phjExcelCellRangeName = phjExcelCellRangeName,
                                                                     phjMaxAttempts = phjAllowedAttempts,
                                                                     phjPrintResults = False)
            if phjTempCellRange is None:
                
                if phjPrintResults == True:
                    print('\nCell range could not be retrieved from Excel workbook.\n')
                
                phjTempDF = None
            
            else:
                # Get name of worksheet and cell co-ordintes from named cell range
                phjTempDestinations = phjTempCellRange.destinations
                
                for phjTempWorksheetName,phjTempCellRangeName in phjTempDestinations:
                    print('Sheet = ',phjTempWorksheetName)
                    print('Range = ',phjTempCellRangeName)
                
                phjTempSheet = phjTempWorkbook[phjTempWorksheetName]
                phjTempRange = phjTempSheet[phjTempCellRangeName]
                
                # Define temporary list to store data...
                phjTempImportedData = []
                
                # Step through each row in cell range.
                # (N.B. Cells returned by iter_rows() are not regular openpyxl.cell.cell.Cell but openpyxl.cell.read_only.ReadOnlyCell.)
                for phjTempRow in phjTempRange:
                    
                    phjTempRowData = phjReadCells(phjTempRow,
                                                  phjDatetimeFormat = phjDatetimeFormat,
                                                  phjMissingValue = phjMissingValue)
                    
                    phjTempImportedData.append(tuple(phjTempRowData))
                
                # Deal with header row
                phjTempVariableNames = phjDealWithHeaderRow(phjTempImportedData,
                                                            phjHeaderRow = phjHeaderRow,
                                                            phjPrintResults = phjPrintResults)
                
                # Convert dataset to pandas DataFrame.
                # Each column now headed with original column headers as seen in Excel file
                # if header row present or with generic labels of 'var1', 'var2', etc. if no
                # header row present.
                phjTempDF = pd.DataFrame(phjTempImportedData, columns=phjTempVariableNames)
    
    
    if phjPrintResults == True:
        print("\nImported data")
        print("-------------")
        print(phjTempDF)
        print('\n')
    
    return phjTempDF



def phjLoadExcelWorkbook(phjExcelPathAndFileName,
                         phjExcelCellRangeName,
                         phjMaxAttempts = 3,
                         phjPrintResults = False):
    
    # Load phjTempWorkbook
    # --------------------
    # If unsuccessful, throws an InvalidFileException (imported in header) and returns None
    
    for i in range(phjMaxAttempts):
        # If necessary, get the path and name of the file
        if (phjExcelPathAndFileName == None) or (i > 0):
            phjExcelPathAndFileName = input('Enter path and filename of Excel workbook: ')
            
        try:
            phjTempWorkbook = openpyxl.load_workbook(filename = phjExcelPathAndFileName, read_only = True, data_only = True)
            
            phjTempCellRange = phjLoadExcelNamedCellRange(phjExcelFileNameOnly = os.path.basename(phjExcelPathAndFileName),
                                                          phjExcelWorkbook = phjTempWorkbook,
                                                          phjExcelCellRangeName = phjExcelCellRangeName,
                                                          phjMaxAttempts = phjMaxAttempts,
                                                          phjPrintResults = False)
            
            break
            
        except FileNotFoundError as e:
            print("\nA FileNotFoundError occurred.\nError number {0}: {1}. File named {2} does not exist at that location.".format(e.args[0],e.args[1],os.path.basename(phjExcelPathAndFileName)))
            
            if i < (phjMaxAttempts-1):
                print("\nPlease re-enter path and filename details.\n")    # Only say 'Please try again' if not last attempt.
            
            else:
                # If workbook can't be found then both phjTempWorkbook and phjTempCellRange are set to None
                print('\nFailed to find Excel workbook after {0} attempts.\n'.format(i+1))
                phjTempWorkbook = None
                phjTempCellRange = None
    
    
    return [phjTempWorkbook,phjTempCellRange]



def phjLoadExcelNamedCellRange(phjExcelFileNameOnly,
                               phjExcelWorkbook,
                               phjExcelCellRangeName,
                               phjMaxAttempts = 3,
                               phjPrintResults = False):
    
    # Get named range of cells
    # ------------------------
    # Function returns None if named_range not found...
    
    for j in range(phjMaxAttempts):
    
        if (phjExcelCellRangeName is None) or (j > 0):
            phjExcelCellRangeName = input('Enter name of cell range in Excel workbook: ')
        
        phjTempCellRange = phjExcelWorkbook.get_named_range(phjExcelCellRangeName)
        
        if phjTempCellRange is not None:
            break
            
        else:
            print("Cell range named {0} was not found in {1}.".format(phjExcelCellRangeName,phjExcelFileNameOnly))
        
            if j < (phjMaxAttempts-1):
                print("\nPlease re-enter path and filename details.\n")    # Don't say 'Please try again' if last attempt has failed.
            
            else:
                print('\nFailed to find named cell range after {0} attempts.'.format(j+1))
            
    return phjTempCellRange



def phjReadCells(phjTempRow,
                 phjDatetimeFormat = "%Y-%m-%d %H:%M:%S",
                 phjMissingValue = "missing",):
    
    # Define temporary list to store values from phjTempCells in phjTempRows
    phjTempData=[]
    
    # Step through each phjTempCell in phjTempRow...
    for phjTempCell in phjTempRow:
        if phjTempCell.value == None:
            phjTempData.append(phjMissingValue)
            
        else:
            if phjTempCell.is_date:
                # If the phjTempCell contains a date, the format of phjTempCell.value is,
                # for example, datetime.datetime(2011, 1, 1, 0, 0). Therefore, reformat
                # using required format.
                phjTempData.append(phjTempCell.value.strftime(phjDatetimeFormat))
            
            elif phjTempCell.data_type == 's':                # TYPE_STRING = 's' AND TYPE_STRING_NULL = 's'
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'f':                # TYPE_FORMULA = 'f'
                # Including 'data_only=True' in openpyxl.load_phjTempWorkbook() means that formulae aren't recognised as formulae, only the resulting value.
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'n':                # TYPE_NUMERIC = 'n'
                # phjTempData.append(Decimal(phjTempCell.internal_value).quantize(Decimal('1.00')))
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'b':                # TYPE_BOOL = 'b'
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'inlineStr':        # TYPE_INLINE = 'inlineStr'
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'e':                # TYPE_ERROR = 'e'
                phjTempData.append(phjTempCell.value)
            
            elif phjTempCell.data_type == 'str':            # TYPE_FORMULA_CACHE_STRING = 'str'
                phjTempData.append(phjTempCell.value)
            
            else:
                phjTempData.append(phjTempCell.value)
                
                
    return phjTempData



def phjDealWithHeaderRow(phjData,
                         phjHeaderRow = False,
                         phjPrintResults = False):
    # This function gets the variable names from the first row of data and
    # removes the header row from the data list. The data is passed by
    # references and, therefore, it can be mutated without having to make
    # a copy of the whole dataset.
    
    # Deal with headers in first phjTempRow...
    if phjHeaderRow:
        # Identify variable names from first phjTempRow of data...
        phjTempVariableNames = phjData[0]
        
        # Remove header names from data...
        del phjData[0]
        
        if phjPrintResults == True:
            print("\nFirst row (containing variable names) has been removed from the data.")
        
    else:
        # If first row doesn't contain variable names then create list of names var1, var2, etc.
        phjTempVariableNames = []
        for i in range (len(phjData[0])):
            phjTempVariableNames.append('var'+str(i+1))
    
    if phjPrintResults == True:
        # Print variable names - just for reference...
        for i in range (len(phjTempVariableNames)):
            print("var",i+1,": ",phjTempVariableNames[i])
    
    return phjTempVariableNames



def phjLoadExcelWorkbookNewOpenpyxl(phjExcelPathAndFileName,
                                    phjExcelCellRangeName,
                                    phjMaxAttempts = 3,
                                    phjPrintResults = False):
    
    # Load phjTempWorkbook
    # --------------------
    # If unsuccessful, throws an InvalidFileException (imported in header) and returns None
    
    for i in range(phjMaxAttempts):
        # If necessary, get the path and name of the file
        if (phjExcelPathAndFileName == None) or (i > 0):
            phjExcelPathAndFileName = input('Enter path and filename of Excel workbook: ')
            
        try:
            phjTempWorkbook = openpyxl.load_workbook(filename = phjExcelPathAndFileName, read_only = True, data_only = True)
            
            break
            
        except FileNotFoundError as e:
            print("\nA FileNotFoundError occurred.\nError number {0}: {1}. File named {2} does not exist at that location.".format(e.args[0],e.args[1],os.path.basename(phjExcelPathAndFileName)))
            
            if i < (phjMaxAttempts-1):
                print("\nPlease re-enter path and filename details.\n")    # Only say 'Please try again' if not last attempt.
            
            else:
                # If workbook can't be found then both phjTempWorkbook and phjTempCellRange are set to None
                print('\nFailed to find Excel workbook after {0} attempts.\n'.format(i+1))
                phjTempWorkbook = None
    
    
    return phjTempWorkbook



def phjLoadExcelNamedCellRangeNewOpenpyxl(phjExcelFileNameOnly,
                                          phjExcelWorkbook,
                                          phjExcelCellRangeName,
                                          phjMaxAttempts = 3,
                                          phjPrintResults = False):
    
    # Get named range of cells
    # ------------------------
    # Function returns None if named_range not found...
    
    for j in range(phjMaxAttempts):
        
        if (phjExcelCellRangeName is None) or (j > 0):
            phjExcelCellRangeName = input('Enter name of cell range in Excel workbook: ')
        
        try:
            phjTempCellRange = phjExcelWorkbook.defined_names[phjExcelCellRangeName]
            
            break
        
        except KeyError as e:
            print("\nA KeyError occurred ({0}). Cell range named '{1}' was not found in workbook '{2}'.".format(e,phjExcelCellRangeName,phjExcelFileNameOnly))
            
            if j < (phjMaxAttempts-1):
                print("\nPlease re-enter cell range name.\n")    # Don't say 'Please try again' if last attempt has failed.
            
            else:
                print('\nFailed to find named cell range after {0} attempts.'.format(j+1))
                phjTempCellRange = None
                
    return phjTempCellRange



if __name__ == '__main__':
    main()